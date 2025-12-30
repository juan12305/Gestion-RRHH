from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook
from datetime import datetime
from trabajadores.models import Trabajador
from contratacion.models import Contratacion
from ingreso.models import Ingreso
from retiro.models import Retiro
from seguridad_social.models import SeguridadSocial
from proyectos.models import Proyecto
from cronograma.models import Cronograma
import os


class Command(BaseCommand):
    help = 'Importa trabajadores desde un archivo Excel'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            help='Ruta al archivo Excel a importar',
            default='excel/1. FORMATO RELACION DE PERSONAL_OCTUBRE.xlsx'
        )
        parser.add_argument(
            '--anio',
            type=int,
            help='Año de los datos a importar (ej: 2024, 2025)',
            default=2025
        )
        parser.add_argument(
            '--sheet',
            type=str,
            help='Nombre de la hoja a importar (ej: "NOVEDADES 2024"). Si no se especifica, se auto-detecta',
            default=None
        )

    def handle(self, *args, **options):
        file_path = options['file']
        anio = options['anio']
        sheet_name = options['sheet']

        # Verificar que el archivo existe
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'El archivo {file_path} no existe'))
            return

        self.stdout.write(self.style.SUCCESS(f'Leyendo archivo: {file_path}'))
        self.stdout.write(self.style.SUCCESS(f'Año a importar: {anio}'))

        try:
            wb = load_workbook(file_path)

            # Auto-detectar la hoja si no se especificó
            if not sheet_name:
                # Buscar hojas con el patrón "NOVEDADES {año}"
                possible_names = [
                    f'NOVEDADES {anio}',
                    f'NOVEDADES {anio} (2)',
                    f'NOVEDADES {str(anio)[-2:]}',  # Ej: "NOVEDADES 24"
                ]

                for name in possible_names:
                    if name in wb.sheetnames:
                        sheet_name = name
                        break

                # Si no se encontró, usar la primera hoja
                if not sheet_name:
                    sheet_name = wb.sheetnames[0]
                    self.stdout.write(self.style.WARNING(f'No se encontró hoja para {anio}, usando: {sheet_name}'))

            # Cargar la hoja específica
            if sheet_name not in wb.sheetnames:
                self.stdout.write(self.style.ERROR(f'La hoja "{sheet_name}" no existe en el archivo'))
                self.stdout.write(f'Hojas disponibles: {", ".join(wb.sheetnames)}')
                return

            ws = wb[sheet_name]
            self.stdout.write(self.style.SUCCESS(f'Usando hoja: {sheet_name}'))

            # Encontrar la fila donde comienzan los datos (después de los headers)
            data_start_row = None
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10), start=1):
                cell_value = str(row[0].value).strip() if row[0].value else ""
                # Buscar la fila que tiene "N°" o un número
                if cell_value and (cell_value == "N°" or cell_value.isdigit()):
                    if cell_value == "N°":
                        data_start_row = row_idx + 1
                    else:
                        data_start_row = row_idx
                    break

            if not data_start_row:
                self.stdout.write(self.style.ERROR('No se pudo encontrar el inicio de los datos'))
                return

            self.stdout.write(f'Los datos comienzan en la fila {data_start_row}')

            # Leer los headers para mapear las columnas
            header_row_idx = data_start_row - 1 if data_start_row > 1 else 1

            # Procesar cada fila de datos
            trabajadores_creados = 0
            trabajadores_actualizados = 0
            errores = 0

            for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row), start=data_start_row):
                try:
                    # Extraer valores de las celdas
                    numero = self._get_cell_value(row, 0)

                    # Si no hay número, saltar esta fila
                    if not numero:
                        continue

                    self.stdout.write(f'\nProcesando trabajador N° {numero}...')

                    with transaction.atomic():
                        # Datos de Identificación (columnas corregidas según el Excel real)
                        fecha_nac = self._parse_date(self._get_cell_value(row, 4))  # Col 4: Fecha de nacimiento
                        fecha_exp = self._parse_date(self._get_cell_value(row, 3))  # Col 3: Fecha expedición cédula

                        # Si no hay fecha de expedición, usar una estimada (18 años después de nacimiento)
                        if not fecha_exp and fecha_nac:
                            from datetime import timedelta
                            fecha_exp = fecha_nac.replace(year=fecha_nac.year + 18) if fecha_nac.year + 18 <= 2025 else fecha_nac
                        elif not fecha_exp and not fecha_nac:
                            # Si ambas son None, usar una fecha por defecto
                            fecha_exp = datetime(2000, 1, 1).date()
                            fecha_nac = datetime(1982, 1, 1).date()
                        elif fecha_exp and not fecha_nac:
                            # Si solo hay fecha de expedición, estimar nacimiento
                            fecha_nac = fecha_exp.replace(year=fecha_exp.year - 18)

                        trabajador_data = {
                            'numero': self._get_cell_value(row, 2),  # Col 2: Número de identificación
                            'tipo': self._map_tipo_identificacion(self._get_cell_value(row, 1)),  # Col 1: Tipo (CC, CE, etc)
                            'primer_apellido': self._get_cell_value(row, 5),  # Col 5: Primer apellido
                            'segundo_apellido': self._get_cell_value(row, 6),  # Col 6: Segundo apellido
                            'primer_nombre': self._get_cell_value(row, 7),  # Col 7: Primer nombre
                            'segundo_nombre': self._get_cell_value(row, 8),  # Col 8: Segundo nombre
                            'fecha_nacimiento': fecha_nac,
                            'fecha_expedicion_cedula': fecha_exp,
                        }

                        # Crear trabajador (sin verificar duplicados)
                        trabajador = Trabajador.objects.create(**trabajador_data)
                        trabajadores_creados += 1
                        self.stdout.write(self.style.SUCCESS(f'  [+] Trabajador creado: {trabajador.nombre_completo}'))

                        # Datos de Contratación (columnas 9-14)
                        tipo_contrato_raw = self._get_cell_value(row, 9)
                        tipo_contrato_map = {
                            'PRESTACION DE SERVICIOS': 'PRESTACION_SERVICIOS',
                            'PRESTACIÓN DE SERVICIOS': 'PRESTACION_SERVICIOS',
                            'TERMINO INDEFINIDO': 'TERMINO_INDEFINIDO',
                            'TÉRMINO INDEFINIDO': 'TERMINO_INDEFINIDO',
                            'TERMINO FIJO': 'TERMINO_FIJO',
                            'TÉRMINO FIJO': 'TERMINO_FIJO',
                            'OBRA O LABOR': 'OBRA_LABOR',
                            'APRENDIZAJE': 'APRENDIZAJE',
                        }
                        tipo_contrato = tipo_contrato_map.get(tipo_contrato_raw.upper(), 'PRESTACION_SERVICIOS') if tipo_contrato_raw else 'PRESTACION_SERVICIOS'

                        cargo = self._get_cell_value(row, 10)
                        salario_contratado = self._parse_decimal(self._get_cell_value(row, 11))
                        municipio_base = self._get_cell_value(row, 12)
                        fecha_inicio_contrato = self._parse_date(self._get_cell_value(row, 13))
                        fecha_final_contrato = self._parse_date(self._get_cell_value(row, 14))

                        # Crear contratacion SIEMPRE (aunque esté vacía)
                        Contratacion.objects.update_or_create(
                            trabajador=trabajador,
                            anio=anio,
                            defaults={
                                'tipo_contrato': tipo_contrato,
                                'cargo': cargo if cargo else '',
                                'salario_contratado': salario_contratado if salario_contratado else 0,
                                'municipio_base': municipio_base.strip().upper() if municipio_base else '',
                                'fecha_inicio_contrato': fecha_inicio_contrato,
                                'fecha_final_contrato': fecha_final_contrato,
                            }
                        )
                        self.stdout.write('  [+] Contratacion guardada')

                        # Datos de Ingreso (columnas 15-18)
                        fecha_ingreso = self._parse_date(self._get_cell_value(row, 15))
                        examen_ingreso = self._parse_date(self._get_cell_value(row, 16))
                        fecha_entrega_epp = self._parse_date(self._get_cell_value(row, 17))
                        fecha_entrega_dotacion = self._parse_date(self._get_cell_value(row, 18))

                        # Crear ingreso SIEMPRE (aunque esté vacío)
                        ing, created = Ingreso.objects.update_or_create(
                            trabajador=trabajador,
                            anio=anio,
                            defaults={
                                'fecha_ingreso': fecha_ingreso,
                                'examen_ingreso': examen_ingreso,
                                'fecha_entrega_epp': fecha_entrega_epp,
                                'fecha_entrega_dotacion': fecha_entrega_dotacion,
                            }
                        )
                        self.stdout.write(f'  [+] Ingreso guardado')

                        # Datos de Retiro (columnas 19-22)
                        fecha_retiro = self._parse_date(self._get_cell_value(row, 19))
                        fecha_liquidacion = self._parse_date(self._get_cell_value(row, 20))
                        valor_liquidacion = self._parse_decimal(self._get_cell_value(row, 21))
                        fecha_examen_retiro = self._parse_date(self._get_cell_value(row, 22))

                        # Crear retiro SIEMPRE (aunque esté vacío)
                        ret, created = Retiro.objects.update_or_create(
                            trabajador=trabajador,
                            anio=anio,
                            defaults={
                                'fecha_retiro': fecha_retiro,
                                'fecha_liquidacion': fecha_liquidacion,
                                'valor_liquidacion': valor_liquidacion,
                                'fecha_examen_retiro': fecha_examen_retiro,
                            }
                        )
                        self.stdout.write(f'  [+] Retiro guardado')

                        # Datos de Seguridad Social (columnas 23-29)
                        eps = self._get_cell_value(row, 23)
                        fecha_afiliacion_eps = self._parse_date(self._get_cell_value(row, 24))
                        caja_compensacion = self._get_cell_value(row, 25)
                        fecha_afiliacion_caja = self._parse_date(self._get_cell_value(row, 26))
                        fondo_pension = self._get_cell_value(row, 27)
                        fecha_afiliacion_pension = self._parse_date(self._get_cell_value(row, 28))
                        arl = self._get_cell_value(row, 29)

                        # Crear seguridad social SIEMPRE (aunque esté vacía)
                        SeguridadSocial.objects.update_or_create(
                            trabajador=trabajador,
                            anio=anio,
                            defaults={
                                'eps': eps if eps else '',
                                'fecha_afiliacion_eps': fecha_afiliacion_eps,
                                'caja_compensacion': caja_compensacion if caja_compensacion else '',
                                'fecha_afiliacion_caja': fecha_afiliacion_caja,
                                'fondo_pension': fondo_pension if fondo_pension else '',
                                'fecha_afiliacion_pension': fecha_afiliacion_pension,
                                'arl': arl.upper() if arl else '',
                            }
                        )
                        self.stdout.write('  [+] Seguridad Social guardada')

                        # Datos de Proyecto (columnas 32-36)
                        proyecto_data = {
                            'administrativo': self._parse_bool(self._get_cell_value(row, 32)),
                            'construccion_instalaciones': self._parse_bool(self._get_cell_value(row, 33)),
                            'construccion_redes': self._parse_bool(self._get_cell_value(row, 34)),
                            'servicios': self._parse_bool(self._get_cell_value(row, 35)),
                            'mantenimiento_redes': self._parse_bool(self._get_cell_value(row, 36)),
                        }

                        Proyecto.objects.update_or_create(
                            trabajador=trabajador,
                            anio=anio,
                            defaults=proyecto_data
                        )
                        self.stdout.write('  [+] Proyecto guardado')

                        # Datos de Cronograma - TODOS LOS 12 MESES
                        # Los meses están en columnas 37-84 (4 columnas por mes × 12 meses = 48 columnas)
                        meses_config = [
                            {'mes': datetime(anio, 1, 1).date(), 'col': 37},   # Enero
                            {'mes': datetime(anio, 2, 1).date(), 'col': 41},   # Febrero
                            {'mes': datetime(anio, 3, 1).date(), 'col': 45},   # Marzo
                            {'mes': datetime(anio, 4, 1).date(), 'col': 49},   # Abril
                            {'mes': datetime(anio, 5, 1).date(), 'col': 53},   # Mayo
                            {'mes': datetime(anio, 6, 1).date(), 'col': 57},   # Junio
                            {'mes': datetime(anio, 7, 1).date(), 'col': 61},   # Julio
                            {'mes': datetime(anio, 8, 1).date(), 'col': 65},   # Agosto
                            {'mes': datetime(anio, 9, 1).date(), 'col': 69},   # Septiembre
                            {'mes': datetime(anio, 10, 1).date(), 'col': 73},  # Octubre
                            {'mes': datetime(anio, 11, 1).date(), 'col': 77},  # Noviembre
                            {'mes': datetime(anio, 12, 1).date(), 'col': 81},  # Diciembre
                        ]

                        cronogramas_creados = 0
                        for mes_info in meses_config:
                            mes_fecha = mes_info['mes']
                            col_base = mes_info['col']

                            # Leer datos del mes (4 columnas: municipio, salario, dias, sueldo)
                            municipio = self._get_cell_value(row, col_base)
                            salario = self._parse_decimal(self._get_cell_value(row, col_base + 1))
                            dias = self._parse_int(self._get_cell_value(row, col_base + 2))
                            sueldo = self._parse_decimal(self._get_cell_value(row, col_base + 3))

                            # Limpiar municipio (convertir N/A y X a None)
                            if municipio and municipio.upper() in ['N/A', 'X']:
                                municipio = None

                            # SIEMPRE crear cronograma (incluso si está vacío) - 141 por mes
                            try:
                                Cronograma.objects.update_or_create(
                                    trabajador=trabajador,
                                    mes=mes_fecha,
                                    defaults={
                                        'municipio_ejecucion': municipio.upper() if municipio else '',
                                        'salario_cotizacion': salario if salario else 0,
                                        'dias_laborados': dias if dias else 0,
                                        'sueldo_devengado': sueldo if sueldo else 0,
                                    }
                                )
                                cronogramas_creados += 1
                            except Exception as crono_error:
                                self.stdout.write(self.style.WARNING(f'    [!] Error en cronograma {mes_fecha}: {str(crono_error)}'))

                        self.stdout.write(f'  [+] {cronogramas_creados} cronograma(s) guardado(s) (12 meses)')

                except Exception as e:
                    errores += 1
                    self.stdout.write(self.style.ERROR(f'  [X] Error en fila {row_idx}: {str(e)}'))

            # Resumen final
            self.stdout.write('\n' + '='*60)
            self.stdout.write(self.style.SUCCESS(f'\n[OK] Importacion completada!'))
            self.stdout.write(f'  - Trabajadores creados: {trabajadores_creados}')
            self.stdout.write(f'  - Trabajadores actualizados: {trabajadores_actualizados}')
            if errores > 0:
                self.stdout.write(self.style.ERROR(f'  - Errores: {errores}'))
            self.stdout.write('='*60 + '\n')

        except Exception as e:
            self.stdout.write(self.style.ERROR(f'\nError al procesar el archivo: {str(e)}'))
            import traceback
            traceback.print_exc()

    def _get_cell_value(self, row, index, default=''):
        """Obtiene el valor de una celda de forma segura"""
        try:
            if index < len(row):
                value = row[index].value
                if value is None:
                    return default
                # No convertir datetime a string, devolverlo tal cual
                if isinstance(value, datetime):
                    return value
                return str(value).strip() if value else default
            return default
        except:
            return default

    def _parse_date(self, value):
        """Convierte un valor a fecha"""
        if not value or value == 'N/A':
            return None

        try:
            # Si ya es un objeto datetime
            if isinstance(value, datetime):
                return value.date()

            # Intentar parsear string
            if isinstance(value, str):
                # Formatos comunes
                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d']:
                    try:
                        return datetime.strptime(value, fmt).date()
                    except:
                        continue
        except:
            pass

        return None

    def _parse_decimal(self, value):
        """Convierte un valor a decimal"""
        if not value or value == 'N/A':
            return None

        try:
            # Si ya es un número (int o float), devolverlo directamente
            if isinstance(value, (int, float)):
                return float(value)

            # Si es string, limpiar y convertir
            if isinstance(value, str):
                # Quitar símbolos de moneda y espacios
                value = value.replace('$', '').strip()

                # Formato colombiano: 1.234.567,89 -> convertir a 1234567.89
                if ',' in value and '.' in value:
                    # Tiene ambos: el punto es separador de miles, la coma es decimal
                    value = value.replace('.', '').replace(',', '.')
                elif ',' in value:
                    # Solo coma: es el decimal
                    value = value.replace(',', '.')
                # Si solo tiene punto, dejarlo así (es el decimal)

            return float(value)
        except:
            return None

    def _parse_int(self, value):
        """Convierte un valor a entero"""
        if not value or value == 'N/A':
            return None

        try:
            return int(float(str(value)))
        except:
            return None

    def _parse_bool(self, value):
        """Convierte un valor a booleano"""
        if not value:
            return False

        value_str = str(value).upper().strip()
        return value_str in ['SÍ', 'SI', 'YES', 'TRUE', '1', 'X', '✓']

    def _map_tipo_identificacion(self, value):
        """Mapea el tipo de identificación del formato completo al código"""
        if not value:
            return 'CC'  # Default

        value_upper = str(value).upper().strip()

        # Mapeo de textos completos a códigos
        mapping = {
            'CÉDULA DE CIUDADANÍA': 'CC',
            'CEDULA DE CIUDADANIA': 'CC',
            'CÉDULA CIUDADANÍA': 'CC',
            'CEDULA': 'CC',
            'CC': 'CC',
            'CÉDULA DE EXTRANJERÍA': 'CE',
            'CEDULA DE EXTRANJERIA': 'CE',
            'CE': 'CE',
            'PASAPORTE': 'PA',
            'PA': 'PA',
            'TARJETA DE IDENTIDAD': 'TI',
            'TI': 'TI',
        }

        return mapping.get(value_upper, 'CC')
