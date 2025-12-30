from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from trabajadores.models import Trabajador
from contratacion.models import Contratacion
from ingreso.models import Ingreso
from retiro.models import Retiro
from seguridad_social.models import SeguridadSocial
from proyectos.models import Proyecto
from cronograma.models import Cronograma
import os


class Command(BaseCommand):
    help = 'Exporta trabajadores a un archivo Excel con el mismo formato que el original'

    def add_arguments(self, parser):
        parser.add_argument(
            '--output',
            type=str,
            help='Nombre del archivo de salida',
            default='excel/1. FORMATO RELACION DE PERSONAL_OCTUBRE.xlsx'
        )
        parser.add_argument(
            '--template',
            type=str,
            help='Archivo plantilla para copiar formato',
            default='excel/1. FORMATO RELACION DE PERSONAL_OCTUBRE.xlsx'
        )
        parser.add_argument(
            '--anio',
            type=int,
            help='Año de los datos a exportar (ej: 2024, 2025)',
            default=2025
        )
        parser.add_argument(
            '--sheet',
            type=str,
            help='Nombre de la hoja a crear/actualizar (ej: "NOVEDADES 2024"). Si no se especifica, se auto-genera',
            default=None
        )

    def handle(self, *args, **options):
        output_path = options['output']
        template_path = options['template']
        anio = options['anio']
        sheet_name = options['sheet']

        # Auto-generar nombre de hoja si no se especificó
        if not sheet_name:
            sheet_name = f'NOVEDADES {anio}'

        # Crear directorio de salida si no existe
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else 'export', exist_ok=True)

        self.stdout.write(self.style.SUCCESS(f'Exportando datos a: {output_path}'))
        self.stdout.write(f'Usando plantilla: {template_path}')
        self.stdout.write(self.style.SUCCESS(f'Año a exportar: {anio}'))
        self.stdout.write(self.style.SUCCESS(f'Hoja a crear/actualizar: {sheet_name}'))

        try:
            # Si el archivo de salida ya existe, usarlo (para preservar otras hojas)
            # Si no, cargar la plantilla
            if os.path.exists(output_path):
                wb = load_workbook(output_path)
                self.stdout.write('  [+] Archivo existente cargado')

                # Si la hoja ya existe, eliminarla para recrearla
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
                    self.stdout.write(f'  [+] Hoja "{sheet_name}" eliminada (será recreada)')
            else:
                # Cargar plantilla para copiar formato
                if os.path.exists(template_path):
                    wb = load_workbook(template_path)
                    self.stdout.write('  [+] Plantilla cargada')
                else:
                    self.stdout.write(self.style.ERROR(f'Plantilla no encontrada: {template_path}'))
                    return

            # Buscar una hoja de plantilla para copiar formato
            template_sheet = None
            template_sheet_names = ['NOVEDADES 2025 (2)', 'NOVEDADES 2025', 'NOVEDADES 2024']
            for ts_name in template_sheet_names:
                if ts_name in wb.sheetnames:
                    template_sheet = wb[ts_name]
                    break

            if not template_sheet:
                template_sheet = wb.active

            # Crear nueva hoja copiando la plantilla
            ws = wb.copy_worksheet(template_sheet)
            ws.title = sheet_name
            self.stdout.write(f'  [+] Hoja "{sheet_name}" creada')

            # Limpiar datos existentes (mantener solo headers)
            if ws.max_row > 4:
                ws.delete_rows(5, ws.max_row - 4)

            # Obtener solo los trabajadores que tienen contratación para el año específico
            contrataciones_ids = Contratacion.objects.filter(anio=anio).values_list('trabajador_id', flat=True)
            trabajadores = Trabajador.objects.filter(id__in=contrataciones_ids).order_by('id')
            total = trabajadores.count()

            self.stdout.write(f'\n  Exportando {total} trabajadores con datos del año {anio}...\n')

            # Escribir datos empezando en la fila 5 (los headers están en 3-4)
            row_idx = 5
            count = 0

            for trabajador in trabajadores:
                try:
                    # Columna 0: Número secuencial
                    ws.cell(row=row_idx, column=1, value=count + 1)

                    # Columnas 1-8: Identificación
                    ws.cell(row=row_idx, column=2, value=trabajador.tipo)
                    ws.cell(row=row_idx, column=3, value=trabajador.numero)
                    ws.cell(row=row_idx, column=4, value=trabajador.fecha_expedicion_cedula)
                    ws.cell(row=row_idx, column=5, value=trabajador.fecha_nacimiento)
                    ws.cell(row=row_idx, column=6, value=trabajador.primer_apellido)
                    ws.cell(row=row_idx, column=7, value=trabajador.segundo_apellido)
                    ws.cell(row=row_idx, column=8, value=trabajador.primer_nombre)
                    ws.cell(row=row_idx, column=9, value=trabajador.segundo_nombre)

                    # Columnas 9-14: Contratación
                    try:
                        contratacion = trabajador.contrataciones.get(anio=anio)
                        ws.cell(row=row_idx, column=10, value=contratacion.get_tipo_contrato_display())
                        ws.cell(row=row_idx, column=11, value=contratacion.cargo)
                        ws.cell(row=row_idx, column=12, value=float(contratacion.salario_contratado))
                        ws.cell(row=row_idx, column=13, value=contratacion.municipio_base)
                        ws.cell(row=row_idx, column=14, value=contratacion.fecha_inicio_contrato)
                        ws.cell(row=row_idx, column=15, value=contratacion.fecha_final_contrato)
                    except Contratacion.DoesNotExist:
                        pass

                    # Columnas 15-18: Ingreso
                    try:
                        ingreso = trabajador.ingresos.get(anio=anio)
                        ws.cell(row=row_idx, column=16, value=ingreso.fecha_ingreso)
                        ws.cell(row=row_idx, column=17, value=ingreso.examen_ingreso)
                        ws.cell(row=row_idx, column=18, value=ingreso.fecha_entrega_epp)
                        ws.cell(row=row_idx, column=19, value=ingreso.fecha_entrega_dotacion)
                    except Ingreso.DoesNotExist:
                        pass

                    # Columnas 19-22: Retiro
                    try:
                        retiro = trabajador.retiros.get(anio=anio)
                        ws.cell(row=row_idx, column=20, value=retiro.fecha_retiro)
                        ws.cell(row=row_idx, column=21, value=retiro.fecha_liquidacion)
                        ws.cell(row=row_idx, column=22, value=float(retiro.valor_liquidacion) if retiro.valor_liquidacion else None)
                        ws.cell(row=row_idx, column=23, value=retiro.fecha_examen_retiro)
                    except Retiro.DoesNotExist:
                        pass

                    # Columnas 23-29: Seguridad Social
                    try:
                        seguridad = trabajador.seguridad_social_registros.get(anio=anio)
                        ws.cell(row=row_idx, column=24, value=seguridad.eps)
                        ws.cell(row=row_idx, column=25, value=seguridad.fecha_afiliacion_eps)
                        ws.cell(row=row_idx, column=26, value=seguridad.caja_compensacion)
                        ws.cell(row=row_idx, column=27, value=seguridad.fecha_afiliacion_caja)
                        ws.cell(row=row_idx, column=28, value=seguridad.fondo_pension)
                        ws.cell(row=row_idx, column=29, value=seguridad.fecha_afiliacion_pension)
                        ws.cell(row=row_idx, column=30, value=seguridad.arl)
                    except SeguridadSocial.DoesNotExist:
                        pass

                    # Columnas 32-36: Proyecto
                    try:
                        proyecto = trabajador.proyectos_asignados.get(anio=anio)
                        ws.cell(row=row_idx, column=33, value='X' if proyecto.administrativo else '')
                        ws.cell(row=row_idx, column=34, value='X' if proyecto.construccion_instalaciones else '')
                        ws.cell(row=row_idx, column=35, value='X' if proyecto.construccion_redes else '')
                        ws.cell(row=row_idx, column=36, value='X' if proyecto.servicios else '')
                        ws.cell(row=row_idx, column=37, value='X' if proyecto.mantenimiento_redes else '')
                    except Proyecto.DoesNotExist:
                        pass

                    # Columnas 37-84: Cronogramas (12 meses × 4 columnas)
                    meses_config = [
                        {'mes': datetime(anio, 1, 1).date(), 'col': 38},   # Enero
                        {'mes': datetime(anio, 2, 1).date(), 'col': 42},   # Febrero
                        {'mes': datetime(anio, 3, 1).date(), 'col': 46},   # Marzo
                        {'mes': datetime(anio, 4, 1).date(), 'col': 50},   # Abril
                        {'mes': datetime(anio, 5, 1).date(), 'col': 54},   # Mayo
                        {'mes': datetime(anio, 6, 1).date(), 'col': 58},   # Junio
                        {'mes': datetime(anio, 7, 1).date(), 'col': 62},   # Julio
                        {'mes': datetime(anio, 8, 1).date(), 'col': 66},   # Agosto
                        {'mes': datetime(anio, 9, 1).date(), 'col': 70},   # Septiembre
                        {'mes': datetime(anio, 10, 1).date(), 'col': 74},  # Octubre
                        {'mes': datetime(anio, 11, 1).date(), 'col': 78},  # Noviembre
                        {'mes': datetime(anio, 12, 1).date(), 'col': 82},  # Diciembre
                    ]

                    for mes_info in meses_config:
                        try:
                            cronograma = Cronograma.objects.get(trabajador=trabajador, mes=mes_info['mes'])
                            col_base = mes_info['col']

                            ws.cell(row=row_idx, column=col_base, value=cronograma.municipio_ejecucion or '')
                            ws.cell(row=row_idx, column=col_base + 1, value=float(cronograma.salario_cotizacion) if cronograma.salario_cotizacion else 0)
                            ws.cell(row=row_idx, column=col_base + 2, value=cronograma.dias_laborados or 0)
                            ws.cell(row=row_idx, column=col_base + 3, value=float(cronograma.sueldo_devengado) if cronograma.sueldo_devengado else 0)
                        except Cronograma.DoesNotExist:
                            # Si no existe el cronograma, dejar las celdas vacías
                            pass

                    count += 1
                    row_idx += 1

                    if count % 20 == 0:
                        self.stdout.write(f'  Exportados {count}/{total} trabajadores...')

                except Exception as e:
                    self.stdout.write(self.style.WARNING(f'  [!] Error exportando trabajador ID={trabajador.id}: {str(e)}'))

            # Guardar el archivo
            wb.save(output_path)

            self.stdout.write('\n' + '='*60)
            self.stdout.write(self.style.SUCCESS(f'\n[OK] Exportación completada!'))
            self.stdout.write(f'  - Trabajadores exportados: {count}')
            self.stdout.write(f'  - Archivo guardado en: {output_path}')
            self.stdout.write('='*60 + '\n')

        except Exception as e:
            self.stdout.write(self.style.ERROR(f'\nError al exportar: {str(e)}'))
            import traceback
            traceback.print_exc()
