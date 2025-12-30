from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import Trabajador
from .serializers import TrabajadorSerializer, TrabajadorListSerializer, TrabajadorDetalleSerializer


class TrabajadorViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar trabajadores.
    Proporciona operaciones CRUD completas.
    """
    queryset = Trabajador.objects.all()
    serializer_class = TrabajadorSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['tipo', 'fecha_nacimiento', 'anio']
    search_fields = ['numero', 'primer_nombre', 'segundo_nombre', 'primer_apellido', 'segundo_apellido']
    ordering_fields = ['primer_apellido', 'fecha_nacimiento', 'fecha_creacion']
    ordering = ['primer_apellido']

    def get_queryset(self):
        """Filtrar trabajadores por año si se proporciona el parámetro"""
        queryset = super().get_queryset()
        anio = self.request.query_params.get('anio', None)
        if anio:
            queryset = queryset.filter(anio=int(anio))
        return queryset

    def get_serializer_class(self):
        """Usar serializer apropiado según la acción"""
        if self.action == 'list':
            return TrabajadorDetalleSerializer
        return TrabajadorSerializer

    def get_serializer_context(self):
        """Agregar el año al contexto del serializer"""
        context = super().get_serializer_context()
        # Obtener el año del query param, por defecto 2025
        context['anio'] = int(self.request.query_params.get('anio', 2025))
        return context

    @action(detail=True, methods=['get'])
    def datos_completos(self, request, pk=None):
        """Endpoint para obtener todos los datos del trabajador incluyendo relaciones"""
        trabajador = self.get_object()
        serializer = TrabajadorDetalleSerializer(trabajador)
        return Response(serializer.data)

    @action(detail=True, methods=['get', 'post', 'put', 'patch', 'delete'], url_path='contratacion')
    def contratacion(self, request, pk=None):
        """
        Endpoint para gestionar la contratación del trabajador
        GET: Ver contratación
        POST: Crear contratación
        PUT/PATCH: Actualizar contratación
        DELETE: Eliminar contratación
        """
        from contratacion.models import Contratacion
        from contratacion.serializers import ContratacionSerializer

        trabajador = self.get_object()

        if request.method == 'GET':
            # Ver contratación
            try:
                contratacion_obj = trabajador.contratacion
                serializer = ContratacionSerializer(contratacion_obj)
                return Response(serializer.data)
            except Contratacion.DoesNotExist:
                return Response(
                    {'error': 'Este trabajador no tiene contratación'},
                    status=status.HTTP_404_NOT_FOUND
                )

        elif request.method == 'POST':
            # Verificar si ya tiene contratación
            if hasattr(trabajador, 'contratacion'):
                return Response(
                    {'error': 'Este trabajador ya tiene una contratación'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Crear contratación
            data = request.data.copy()
            data['trabajador'] = trabajador.id
            serializer = ContratacionSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        elif request.method in ['PUT', 'PATCH']:
            # Actualizar contratación existente
            try:
                contratacion_obj = trabajador.contratacion
            except Contratacion.DoesNotExist:
                return Response(
                    {'error': 'Este trabajador no tiene contratación'},
                    status=status.HTTP_404_NOT_FOUND
                )

            data = request.data.copy()
            data['trabajador'] = trabajador.id
            partial = request.method == 'PATCH'
            serializer = ContratacionSerializer(contratacion_obj, data=data, partial=partial)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        elif request.method == 'DELETE':
            # Eliminar contratación
            try:
                contratacion_obj = trabajador.contratacion
                contratacion_obj.delete()
                return Response(status=status.HTTP_204_NO_CONTENT)
            except Contratacion.DoesNotExist:
                return Response(
                    {'error': 'Este trabajador no tiene contratación'},
                    status=status.HTTP_404_NOT_FOUND
                )

    @action(detail=True, methods=['get', 'post', 'put', 'patch', 'delete'], url_path='ingreso')
    def ingreso(self, request, pk=None):
        """
        Endpoint para gestionar el ingreso del trabajador
        GET: Ver ingreso
        POST: Crear ingreso
        PUT/PATCH: Actualizar ingreso
        DELETE: Eliminar ingreso
        """
        from ingreso.models import Ingreso
        from ingreso.serializers import IngresoSerializer

        trabajador = self.get_object()

        if request.method == 'GET':
            # Ver ingreso
            try:
                ingreso_obj = trabajador.ingreso
                serializer = IngresoSerializer(ingreso_obj)
                return Response(serializer.data)
            except Ingreso.DoesNotExist:
                return Response(
                    {'error': 'Este trabajador no tiene ingreso'},
                    status=status.HTTP_404_NOT_FOUND
                )

        elif request.method == 'POST':
            # Verificar si ya tiene ingreso
            if hasattr(trabajador, 'ingreso'):
                return Response(
                    {'error': 'Este trabajador ya tiene un ingreso'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Crear ingreso
            data = request.data.copy()
            data['trabajador'] = trabajador.id
            serializer = IngresoSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        elif request.method in ['PUT', 'PATCH']:
            # Actualizar ingreso existente
            try:
                ingreso_obj = trabajador.ingreso
            except Ingreso.DoesNotExist:
                return Response(
                    {'error': 'Este trabajador no tiene ingreso'},
                    status=status.HTTP_404_NOT_FOUND
                )

            data = request.data.copy()
            data['trabajador'] = trabajador.id
            partial = request.method == 'PATCH'
            serializer = IngresoSerializer(ingreso_obj, data=data, partial=partial)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        elif request.method == 'DELETE':
            # Eliminar ingreso
            try:
                ingreso_obj = trabajador.ingreso
                ingreso_obj.delete()
                return Response(status=status.HTTP_204_NO_CONTENT)
            except Ingreso.DoesNotExist:
                return Response(
                    {'error': 'Este trabajador no tiene ingreso'},
                    status=status.HTTP_404_NOT_FOUND
                )

    @action(detail=True, methods=['get', 'post', 'put', 'patch', 'delete'], url_path='retiro')
    def retiro(self, request, pk=None):
        """
        Endpoint para gestionar el retiro del trabajador
        GET: Ver retiro
        POST: Crear retiro
        PUT/PATCH: Actualizar retiro
        DELETE: Eliminar retiro
        """
        from retiro.models import Retiro
        from retiro.serializers import RetiroSerializer

        trabajador = self.get_object()

        if request.method == 'GET':
            # Ver retiro
            try:
                retiro_obj = trabajador.retiro
                serializer = RetiroSerializer(retiro_obj)
                return Response(serializer.data)
            except Retiro.DoesNotExist:
                return Response(
                    {'error': 'Este trabajador no tiene retiro'},
                    status=status.HTTP_404_NOT_FOUND
                )

        elif request.method == 'POST':
            # Verificar si ya tiene retiro
            if hasattr(trabajador, 'retiro'):
                return Response(
                    {'error': 'Este trabajador ya tiene un retiro'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Crear retiro
            data = request.data.copy()
            data['trabajador'] = trabajador.id
            serializer = RetiroSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        elif request.method in ['PUT', 'PATCH']:
            # Actualizar retiro existente
            try:
                retiro_obj = trabajador.retiro
            except Retiro.DoesNotExist:
                return Response(
                    {'error': 'Este trabajador no tiene retiro'},
                    status=status.HTTP_404_NOT_FOUND
                )

            data = request.data.copy()
            data['trabajador'] = trabajador.id
            partial = request.method == 'PATCH'
            serializer = RetiroSerializer(retiro_obj, data=data, partial=partial)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        elif request.method == 'DELETE':
            # Eliminar retiro
            try:
                retiro_obj = trabajador.retiro
                retiro_obj.delete()
                return Response(status=status.HTTP_204_NO_CONTENT)
            except Retiro.DoesNotExist:
                return Response(
                    {'error': 'Este trabajador no tiene retiro'},
                    status=status.HTTP_404_NOT_FOUND
                )

    @action(detail=True, methods=['get', 'post', 'put', 'patch', 'delete'], url_path='seguridad-social')
    def seguridad_social(self, request, pk=None):
        """
        Endpoint para gestionar la seguridad social del trabajador
        GET: Ver seguridad social
        POST: Crear seguridad social
        PUT/PATCH: Actualizar seguridad social
        DELETE: Eliminar seguridad social
        """
        from seguridad_social.models import SeguridadSocial
        from seguridad_social.serializers import SeguridadSocialSerializer

        trabajador = self.get_object()

        if request.method == 'GET':
            # Ver seguridad social
            try:
                seguridad_social_obj = trabajador.seguridad_social
                serializer = SeguridadSocialSerializer(seguridad_social_obj)
                return Response(serializer.data)
            except SeguridadSocial.DoesNotExist:
                return Response(
                    {'error': 'Este trabajador no tiene seguridad social'},
                    status=status.HTTP_404_NOT_FOUND
                )

        elif request.method == 'POST':
            # Verificar si ya tiene seguridad social
            if hasattr(trabajador, 'seguridad_social'):
                return Response(
                    {'error': 'Este trabajador ya tiene seguridad social'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Crear seguridad social
            data = request.data.copy()
            data['trabajador'] = trabajador.id
            serializer = SeguridadSocialSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        elif request.method in ['PUT', 'PATCH']:
            # Actualizar seguridad social existente
            try:
                seguridad_social_obj = trabajador.seguridad_social
            except SeguridadSocial.DoesNotExist:
                return Response(
                    {'error': 'Este trabajador no tiene seguridad social'},
                    status=status.HTTP_404_NOT_FOUND
                )

            data = request.data.copy()
            data['trabajador'] = trabajador.id
            partial = request.method == 'PATCH'
            serializer = SeguridadSocialSerializer(seguridad_social_obj, data=data, partial=partial)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        elif request.method == 'DELETE':
            # Eliminar seguridad social
            try:
                seguridad_social_obj = trabajador.seguridad_social
                seguridad_social_obj.delete()
                return Response(status=status.HTTP_204_NO_CONTENT)
            except SeguridadSocial.DoesNotExist:
                return Response(
                    {'error': 'Este trabajador no tiene seguridad social'},
                    status=status.HTTP_404_NOT_FOUND
                )

    @action(detail=True, methods=['get', 'post', 'put', 'patch', 'delete'], url_path='proyectos')
    def proyectos(self, request, pk=None):
        """
        Endpoint para gestionar los proyectos del trabajador
        GET: Ver proyectos
        POST: Crear proyectos
        PUT/PATCH: Actualizar proyectos
        DELETE: Eliminar proyectos
        """
        from proyectos.models import Proyecto
        from proyectos.serializers import ProyectoSerializer

        trabajador = self.get_object()

        if request.method == 'GET':
            # Ver proyectos
            try:
                proyectos_obj = trabajador.proyectos
                serializer = ProyectoSerializer(proyectos_obj)
                return Response(serializer.data)
            except Proyecto.DoesNotExist:
                return Response(
                    {'error': 'Este trabajador no tiene proyectos asignados'},
                    status=status.HTTP_404_NOT_FOUND
                )

        elif request.method == 'POST':
            # Verificar si ya tiene proyectos
            if hasattr(trabajador, 'proyectos'):
                return Response(
                    {'error': 'Este trabajador ya tiene proyectos asignados'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Crear proyectos
            data = request.data.copy()
            data['trabajador'] = trabajador.id
            serializer = ProyectoSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        elif request.method in ['PUT', 'PATCH']:
            # Actualizar proyectos existentes
            try:
                proyectos_obj = trabajador.proyectos
            except Proyecto.DoesNotExist:
                return Response(
                    {'error': 'Este trabajador no tiene proyectos asignados'},
                    status=status.HTTP_404_NOT_FOUND
                )

            data = request.data.copy()
            data['trabajador'] = trabajador.id
            partial = request.method == 'PATCH'
            serializer = ProyectoSerializer(proyectos_obj, data=data, partial=partial)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        elif request.method == 'DELETE':
            # Eliminar proyectos
            try:
                proyectos_obj = trabajador.proyectos
                proyectos_obj.delete()
                return Response(status=status.HTTP_204_NO_CONTENT)
            except Proyecto.DoesNotExist:
                return Response(
                    {'error': 'Este trabajador no tiene proyectos asignados'},
                    status=status.HTTP_404_NOT_FOUND
                )

    @action(detail=True, methods=['get', 'post'], url_path='cronograma')
    def cronograma(self, request, pk=None):
        """
        Endpoint para gestionar el cronograma del trabajador
        GET: Listar todos los cronogramas (meses) del trabajador
        POST: Crear un nuevo registro de cronograma para un mes específico
        """
        from cronograma.models import Cronograma
        from cronograma.serializers import CronogramaSerializer

        trabajador = self.get_object()

        if request.method == 'GET':
            # Listar todos los cronogramas del trabajador
            cronogramas = trabajador.cronogramas.all()

            # Filtrar por mes si se proporciona el parámetro
            mes = request.query_params.get('mes', None)
            if mes:
                cronogramas = cronogramas.filter(mes__startswith=mes)

            serializer = CronogramaSerializer(cronogramas, many=True)
            return Response(serializer.data)

        elif request.method == 'POST':
            # Crear un nuevo registro de cronograma
            data = request.data.copy()
            data['trabajador'] = trabajador.id
            serializer = CronogramaSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['put', 'patch', 'delete'], url_path='cronograma/(?P<cronograma_id>[^/.]+)')
    def cronograma_detail(self, request, pk=None, cronograma_id=None):
        """
        Endpoint para actualizar o eliminar un cronograma específico
        PUT/PATCH: Actualizar cronograma
        DELETE: Eliminar cronograma
        """
        from cronograma.models import Cronograma
        from cronograma.serializers import CronogramaSerializer

        trabajador = self.get_object()

        try:
            cronograma_obj = trabajador.cronogramas.get(id=cronograma_id)
        except Cronograma.DoesNotExist:
            return Response(
                {'error': 'Cronograma no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        if request.method in ['PUT', 'PATCH']:
            data = request.data.copy()
            data['trabajador'] = trabajador.id
            partial = request.method == 'PATCH'
            serializer = CronogramaSerializer(cronograma_obj, data=data, partial=partial)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        elif request.method == 'DELETE':
            cronograma_obj.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['get'], url_path='exportar-excel', permission_classes=[AllowAny])
    def exportar_excel(self, request):
        """
        Exporta todos los trabajadores a Excel usando el mismo formato que la plantilla original
        GET /api/trabajadores/exportar-excel/
        """
        from datetime import datetime
        from openpyxl import load_workbook
        from contratacion.models import Contratacion
        from ingreso.models import Ingreso
        from retiro.models import Retiro
        from seguridad_social.models import SeguridadSocial
        from proyectos.models import Proyecto
        from cronograma.models import Cronograma
        import os

        # Ruta de la plantilla
        template_path = 'excel/1. FORMATO RELACION DE PERSONAL_OCTUBRE.xlsx'

        # Verificar que la plantilla existe
        if not os.path.exists(template_path):
            return Response(
                {'error': f'Plantilla no encontrada: {template_path}'},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            # Cargar la plantilla
            wb = load_workbook(template_path)
            ws_template = wb.active

            # Eliminar todas las hojas excepto la plantilla
            sheets_to_remove = [sheet for sheet in wb.worksheets if sheet != ws_template]
            for sheet in sheets_to_remove:
                wb.remove(sheet)

            # Crear dos nuevas hojas para 2024 y 2025
            ws_2024 = wb.create_sheet("NOVEDADES 2024", 0)
            ws_2025 = wb.create_sheet("NOVEDADES 2025", 1)

            # Copiar los headers de la plantilla a ambas hojas (filas 3-4 de plantilla → filas 1-2 de nuevas hojas)
            # Mapeo: fila 3 de plantilla → fila 1, fila 4 de plantilla → fila 2
            row_mapping = {3: 1, 4: 2}

            for template_row, new_row in row_mapping.items():
                for col_idx in range(1, ws_template.max_column + 1):
                    cell_template = ws_template.cell(row=template_row, column=col_idx)

                    # Copiar a hoja 2024
                    cell_2024 = ws_2024.cell(row=new_row, column=col_idx)
                    cell_2024.value = cell_template.value
                    if cell_template.has_style:
                        cell_2024.font = cell_template.font.copy()
                        cell_2024.border = cell_template.border.copy()
                        cell_2024.fill = cell_template.fill.copy()
                        cell_2024.number_format = cell_template.number_format
                        cell_2024.protection = cell_template.protection.copy()
                        cell_2024.alignment = cell_template.alignment.copy()

                    # Copiar a hoja 2025
                    cell_2025 = ws_2025.cell(row=new_row, column=col_idx)
                    cell_2025.value = cell_template.value
                    if cell_template.has_style:
                        cell_2025.font = cell_template.font.copy()
                        cell_2025.border = cell_template.border.copy()
                        cell_2025.fill = cell_template.fill.copy()
                        cell_2025.number_format = cell_template.number_format
                        cell_2025.protection = cell_template.protection.copy()
                        cell_2025.alignment = cell_template.alignment.copy()

            # Copiar merged cells (celdas combinadas) ajustando las filas
            # Solo copiar las de las filas 3 y 4, y ajustarlas a filas 1 y 2
            for merged_cell_range in ws_template.merged_cells.ranges:
                min_row = merged_cell_range.min_row
                max_row = merged_cell_range.max_row

                # Solo copiar si está en las filas 3 o 4 de la plantilla
                if min_row >= 3 and max_row <= 4:
                    # Ajustar las filas: restar 2 para mover de 3-4 a 1-2
                    from openpyxl.utils import get_column_letter
                    min_col_letter = get_column_letter(merged_cell_range.min_col)
                    max_col_letter = get_column_letter(merged_cell_range.max_col)
                    new_min_row = min_row - 2
                    new_max_row = max_row - 2

                    new_range = f"{min_col_letter}{new_min_row}:{max_col_letter}{new_max_row}"

                    # Copiar a hoja 2024
                    ws_2024.merge_cells(new_range)
                    # Copiar a hoja 2025
                    ws_2025.merge_cells(new_range)

            # Copiar anchos de columna
            for col_idx in range(1, ws_template.max_column + 1):
                col_letter = get_column_letter(col_idx)
                ws_2024.column_dimensions[col_letter].width = ws_template.column_dimensions[col_letter].width
                ws_2025.column_dimensions[col_letter].width = ws_template.column_dimensions[col_letter].width

            # Copiar alturas de fila (para los headers)
            # Mapear las alturas de las filas 3-4 de la plantilla a las filas 1-2 de las nuevas hojas
            ws_2024.row_dimensions[1].height = ws_template.row_dimensions[3].height
            ws_2024.row_dimensions[2].height = ws_template.row_dimensions[4].height
            ws_2025.row_dimensions[1].height = ws_template.row_dimensions[3].height
            ws_2025.row_dimensions[2].height = ws_template.row_dimensions[4].height

            # Congelar paneles: columnas A-I (1-9) estáticas, desde J en adelante se mueven
            # También congela las filas 1-2 (headers) para que permanezcan visibles al desplazarse verticalmente
            ws_2024.freeze_panes = 'J3'  # Congela hasta columna I y fila 2
            ws_2025.freeze_panes = 'J3'  # Congela hasta columna I y fila 2

            # Eliminar la hoja de plantilla original
            wb.remove(ws_template)

            # Función para escribir datos de trabajadores en una hoja
            def escribir_trabajadores_en_hoja(ws, trabajadores_filtrados):
                row_idx = 3  # Los datos empiezan en la fila 3 (después de headers en filas 1-2)
                count = 0

                for trabajador in trabajadores_filtrados:
                    try:
                        # Columna 0: Número secuencial
                        ws.cell(row=row_idx, column=1, value=count + 1)

                        # Columnas 1-8: Identificación
                        ws.cell(row=row_idx, column=2, value=trabajador.tipo or '')
                        ws.cell(row=row_idx, column=3, value=trabajador.numero or '')
                        ws.cell(row=row_idx, column=4, value=trabajador.fecha_expedicion_cedula)
                        ws.cell(row=row_idx, column=5, value=trabajador.fecha_nacimiento)
                        ws.cell(row=row_idx, column=6, value=trabajador.primer_apellido or '')
                        ws.cell(row=row_idx, column=7, value=trabajador.segundo_apellido or '')
                        ws.cell(row=row_idx, column=8, value=trabajador.primer_nombre or '')
                        ws.cell(row=row_idx, column=9, value=trabajador.segundo_nombre or '')

                        # Columnas 9-14: Contratación
                        try:
                            contratacion = trabajador.contrataciones.filter(anio=trabajador.anio).first()
                            if contratacion:
                                ws.cell(row=row_idx, column=10, value=contratacion.get_tipo_contrato_display() or '')
                                ws.cell(row=row_idx, column=11, value=contratacion.cargo or '')
                                ws.cell(row=row_idx, column=12, value=float(contratacion.salario_contratado) if contratacion.salario_contratado else 0)
                                ws.cell(row=row_idx, column=13, value=contratacion.municipio_base or '')
                                ws.cell(row=row_idx, column=14, value=contratacion.fecha_inicio_contrato)
                                ws.cell(row=row_idx, column=15, value=contratacion.fecha_final_contrato)
                        except Exception:
                            pass

                        # Columnas 15-18: Ingreso
                        try:
                            ingreso = trabajador.ingresos.filter(anio=trabajador.anio).first()
                            if ingreso:
                                ws.cell(row=row_idx, column=16, value=ingreso.fecha_ingreso)
                                ws.cell(row=row_idx, column=17, value=ingreso.examen_ingreso)
                                ws.cell(row=row_idx, column=18, value=ingreso.fecha_entrega_epp)
                                ws.cell(row=row_idx, column=19, value=ingreso.fecha_entrega_dotacion)
                        except Exception:
                            pass

                        # Columnas 19-22: Retiro
                        try:
                            retiro = trabajador.retiros.filter(anio=trabajador.anio).first()
                            if retiro:
                                ws.cell(row=row_idx, column=20, value=retiro.fecha_retiro)
                                ws.cell(row=row_idx, column=21, value=retiro.fecha_liquidacion)
                                ws.cell(row=row_idx, column=22, value=float(retiro.valor_liquidacion) if retiro.valor_liquidacion else None)
                                ws.cell(row=row_idx, column=23, value=retiro.fecha_examen_retiro)
                        except Exception:
                            pass

                        # Columnas 23-29: Seguridad Social
                        try:
                            seguridad = trabajador.seguridad_social_registros.filter(anio=trabajador.anio).first()
                            if seguridad:
                                ws.cell(row=row_idx, column=24, value=seguridad.eps)
                                ws.cell(row=row_idx, column=25, value=seguridad.fecha_afiliacion_eps)
                                ws.cell(row=row_idx, column=26, value=seguridad.caja_compensacion)
                                ws.cell(row=row_idx, column=27, value=seguridad.fecha_afiliacion_caja)
                                ws.cell(row=row_idx, column=28, value=seguridad.fondo_pension)
                                ws.cell(row=row_idx, column=29, value=seguridad.fecha_afiliacion_pension)
                                ws.cell(row=row_idx, column=30, value=seguridad.arl)
                        except Exception:
                            pass

                        # Columnas 32-36: Proyecto
                        try:
                            proyecto = trabajador.proyectos_asignados.filter(anio=trabajador.anio).first()
                            if proyecto:
                                ws.cell(row=row_idx, column=33, value='X' if proyecto.administrativo else '')
                                ws.cell(row=row_idx, column=34, value='X' if proyecto.construccion_instalaciones else '')
                                ws.cell(row=row_idx, column=35, value='X' if proyecto.construccion_redes else '')
                                ws.cell(row=row_idx, column=36, value='X' if proyecto.servicios else '')
                                ws.cell(row=row_idx, column=37, value='X' if proyecto.mantenimiento_redes else '')
                        except Exception:
                            pass

                        # Columnas 37-84: Cronogramas (12 meses × 4 columnas)
                        # Usar el año del trabajador para los cronogramas
                        anio_trabajador = trabajador.anio
                        meses_config = [
                            {'mes': datetime(anio_trabajador, 1, 1).date(), 'col': 38},   # Enero
                            {'mes': datetime(anio_trabajador, 2, 1).date(), 'col': 42},   # Febrero
                            {'mes': datetime(anio_trabajador, 3, 1).date(), 'col': 46},   # Marzo
                            {'mes': datetime(anio_trabajador, 4, 1).date(), 'col': 50},   # Abril
                            {'mes': datetime(anio_trabajador, 5, 1).date(), 'col': 54},   # Mayo
                            {'mes': datetime(anio_trabajador, 6, 1).date(), 'col': 58},   # Junio
                            {'mes': datetime(anio_trabajador, 7, 1).date(), 'col': 62},   # Julio
                            {'mes': datetime(anio_trabajador, 8, 1).date(), 'col': 66},   # Agosto
                            {'mes': datetime(anio_trabajador, 9, 1).date(), 'col': 70},   # Septiembre
                            {'mes': datetime(anio_trabajador, 10, 1).date(), 'col': 74},  # Octubre
                            {'mes': datetime(anio_trabajador, 11, 1).date(), 'col': 78},  # Noviembre
                            {'mes': datetime(anio_trabajador, 12, 1).date(), 'col': 82},  # Diciembre
                        ]

                        for mes_info in meses_config:
                            try:
                                cronograma = trabajador.cronogramas.filter(mes=mes_info['mes']).first()
                                if cronograma:
                                    col_base = mes_info['col']
                                    ws.cell(row=row_idx, column=col_base, value=cronograma.municipio_ejecucion or '')
                                    ws.cell(row=row_idx, column=col_base + 1, value=float(cronograma.salario_cotizacion) if cronograma.salario_cotizacion else 0)
                                    ws.cell(row=row_idx, column=col_base + 2, value=cronograma.dias_laborados or 0)
                                    ws.cell(row=row_idx, column=col_base + 3, value=float(cronograma.sueldo_devengado) if cronograma.sueldo_devengado else 0)
                            except Exception:
                                # Si no existe el cronograma, dejar las celdas vacías
                                pass

                        count += 1
                        row_idx += 1

                    except Exception as e:
                        # Continuar con el siguiente trabajador si hay error
                        pass

            # Filtrar trabajadores por año
            trabajadores_2024 = Trabajador.objects.filter(anio=2024).order_by('id')
            trabajadores_2025 = Trabajador.objects.filter(anio=2025).order_by('id')

            # Escribir datos en cada hoja
            escribir_trabajadores_en_hoja(ws_2024, trabajadores_2024)
            escribir_trabajadores_en_hoja(ws_2025, trabajadores_2025)

            # Crear la respuesta HTTP con el archivo Excel
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f'RELACION_PERSONAL_EXPORT_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            response['Content-Disposition'] = f'attachment; filename={filename}'

            # Guardar el workbook en la respuesta
            wb.save(response)
            return response

        except Exception as e:
            return Response(
                {'error': f'Error al exportar: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
